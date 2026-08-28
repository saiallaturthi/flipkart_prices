import json
import os
import re
import time
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def get_main_excel_path():
    """Locates the main Flipkart input workbook."""
    possible_paths = [
        "FK Walkthrough 11th aug.xlsx",
        "FK Walkthrough 11th aug",
        r"C:\Users\sai\OneDrive\Desktop\flipkart\FK Walkthrough 11th aug.xlsx",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return "FK Walkthrough 11th aug.xlsx"


def get_clean_price(driver):
    """
    Extracts the exact selling price from the Flipkart product page.
    Uses explicit DOM element waits to avoid fetching strike-through/MRP prices.
    """
    # Primary price container classes on Flipkart
    price_selectors = [
        "div.Nx9bqj.CxhGGd",     # Main active selling price on desktop
        "div.Nx9bqj",             # Generic updated selling price class
        "div._30jeq3._16Jk6d",     # Older main product selling price
        "div._30jeq3",             # Older standard selling price
    ]

    # 1. Selenium Explicit Wait on live DOM element
    for selector in price_selectors:
        try:
            elem = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
            )
            text = elem.text.strip()
            cleaned = re.sub(r"[^\d.]", "", text)
            if cleaned:
                return cleaned
        except Exception:
            continue

    # 2. BeautifulSoup parse scoped strictly to price classes (ignoring MRP/strikethrough classes like .yRaY8j or ._3I9_wc)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    for selector in price_selectors:
        price_elem = soup.select_one(selector)
        if price_elem and price_elem.text:
            cleaned = re.sub(r"[^\d.]", "", price_elem.text.strip())
            if cleaned:
                return cleaned

    # 3. Fallback Method: JSON-LD structured data
    json_ld_scripts = soup.find_all("script", type="application/ld+json")
    for script in json_ld_scripts:
        if script.string:
            try:
                data = json.loads(script.string)
                if isinstance(data, list):
                    data = data[0]
                if "offers" in data:
                    offers = data["offers"]
                    if isinstance(offers, list):
                        offers = offers[0]
                    price = offers.get("price") or offers.get("lowPrice")
                    if price and str(price).strip() not in ["0", "None", ""]:
                        return str(price).replace(",", "").strip()
            except Exception:
                continue

    return "N/A"


def append_issue_row(target_file_name, row_data):
    """Appends discrepancy row to target issue workbook."""
    headers = [
        "Flipkart Serial Number",
        "Seller SKU Id",
        "Product Title",
        "AZ Price",
        "Current Price",
        "Remark",
    ]

    target_path = None
    for path_option in [target_file_name, f"{target_file_name}.xlsx"]:
        if os.path.exists(path_option):
            target_path = path_option
            break

    if not target_path:
        target_path = f"{target_file_name}.xlsx"

    issue_wb = None
    issue_ws = None

    if os.path.exists(target_path):
        try:
            issue_wb = openpyxl.load_workbook(target_path)
            issue_ws = issue_wb.active
        except Exception:
            issue_wb = openpyxl.Workbook()
            issue_ws = issue_wb.active
            issue_ws.append(headers)
    else:
        issue_wb = openpyxl.Workbook()
        issue_ws = issue_wb.active
        issue_ws.append(headers)

    if issue_ws.max_row < 1 or issue_ws.cell(row=1, column=1).value is None:
        issue_ws.append(headers)

    issue_ws.append(row_data)

    try:
        issue_wb.save(target_path)
    except Exception as e:
        print(f" [Error] Could not save issue file '{target_path}': {e}")


def main():
    main_excel_file = get_main_excel_path()
    if not os.path.exists(main_excel_file):
        raise FileNotFoundError(f"Could not locate '{main_excel_file}'")

    wb = openpyxl.load_workbook(main_excel_file, data_only=True)

    # Configure Headless Chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )

    print("Launching Headless Chrome Driver for Flipkart...")
    driver = webdriver.Chrome(options=chrome_options)

    skip_sheets = {"amzon home", "amazon home", "sheet1", "sheet 1"}

    try:
        for sheet_name in wb.sheetnames:
            clean_sheet_name = sheet_name.strip()
            if clean_sheet_name.lower() in skip_sheets:
                print(f"\nSkipping sheet: '{sheet_name}'")
                continue

            ws = wb[sheet_name]
            print(f"\n{'=' * 60}\nProcessing sheet: '{sheet_name}'\n{'=' * 60}")

            headers = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in ws[1]
            ]

            def get_col_idx(name, default_idx=-1):
                return headers.index(name) + 1 if name in headers else default_idx

            url_col_idx = get_col_idx("URL", 6)
            price_col_idx = get_col_idx("Current Price", get_col_idx("Current", -1))
            stock_col_idx = get_col_idx("FK Stock", 13)
            az_price_col_idx = get_col_idx("AZ Price", get_col_idx("Amazon Price", -1))
            fsn_col_idx = get_col_idx("Flipkart Serial Number", get_col_idx("FSN", 1))
            sku_col_idx = get_col_idx("Seller SKU Id", get_col_idx("SKU", 2))
            title_col_idx = get_col_idx("Product Title", get_col_idx("Title", 3))
            category_col_idx = get_col_idx("Category", get_col_idx("Sub-Category", -1))

            if price_col_idx == -1:
                print(f"Skipping sheet '{sheet_name}': 'Current Price' column not found.")
                continue

            total_rows = ws.max_row

            for row in range(2, total_rows + 1):
                # Stock Check >= 50
                stock_val = ws.cell(row=row, column=stock_col_idx).value
                try:
                    stock_count = float(stock_val) if stock_val is not None else 0
                except (ValueError, TypeError):
                    stock_count = 0

                if stock_count < 50:
                    print(f"[{row-1}/{total_rows-1}] Stock is {stock_val} (< 50). Skipping Row {row}.")
                    continue

                url_cell = ws.cell(row=row, column=url_col_idx).value
                url = str(url_cell).strip() if url_cell else ""

                if not url.startswith("http"):
                    print(f"[{row-1}/{total_rows-1}] Invalid/Missing URL. Skipping Row {row}.")
                    continue

                try:
                    driver.get(url)

                    # Extract price directly via explicit waits and specific containers
                    current_price_str = get_clean_price(driver)

                    # Process current scraped price
                    if current_price_str not in ["N/A", "Not Found", ""]:
                        try:
                            current_price_val = (
                                float(current_price_str)
                                if "." in current_price_str
                                else int(current_price_str)
                            )
                            ws.cell(row=row, column=price_col_idx, value=current_price_val)
                            print_price = f"₹{current_price_val}"
                        except ValueError:
                            current_price_val = None
                            ws.cell(row=row, column=price_col_idx, value="N/A")
                            print_price = "N/A"
                    else:
                        current_price_val = None
                        ws.cell(row=row, column=price_col_idx, value="N/A")
                        print_price = "N/A"

                    # Extract and parse AZ Price safely
                    az_price_val = (
                        ws.cell(row=row, column=az_price_col_idx).value
                        if az_price_col_idx != -1
                        else None
                    )
                    az_clean_str = (
                        re.sub(r"[^\d.]", "", str(az_price_val))
                        if az_price_val is not None
                        else ""
                    )

                    # Evaluate difference only when both prices are valid numbers
                    if current_price_val is not None and az_clean_str != "":
                        try:
                            az_num = float(az_clean_str)
                            curr_num = float(current_price_val)
                            price_diff = abs(az_num - curr_num)

                            if price_diff > 10:
                                if curr_num < az_num:
                                    remark = "Current price is Less than Amazon price"
                                else:
                                    remark = "Current price is More than Amazon price"

                                fsn = ws.cell(row=row, column=fsn_col_idx).value or ""
                                sku = ws.cell(row=row, column=sku_col_idx).value or ""
                                title = ws.cell(row=row, column=title_col_idx).value or ""
                                category = str(
                                    ws.cell(row=row, column=category_col_idx).value or ""
                                ).lower()

                                report_row = [fsn, sku, title, az_num, curr_num, remark]

                                target_file = clean_sheet_name

                                # Category splitting for Roma vs Power Banks
                                if "roma" in clean_sheet_name.lower():
                                    if (
                                        "powerbank" in category
                                        or "power bank" in category
                                        or "power bank" in str(title).lower()
                                    ):
                                        target_file = "Power Banks"
                                    else:
                                        target_file = "Roma"

                                append_issue_row(target_file, report_row)
                                print(f"[{row-1}/{total_rows-1}] Row {row} Updated -> Price: {print_price} | Logged to '{target_file}': {remark}")
                            else:
                                print(f"[{row-1}/{total_rows-1}] Row {row} Updated -> Price: {print_price} | Diff <= 10.")

                        except (ValueError, TypeError) as e:
                            print(f"[{row-1}/{total_rows-1}] Row {row} Updated -> Price: {print_price} | Calculation error: {e}")
                    else:
                        print(f"[{row-1}/{total_rows-1}] Row {row} Updated -> Price: {print_price} | Skipping diff check (AZ or Current Price is N/A)")

                except Exception as err:
                    print(f"[{row-1}/{total_rows-1}] Row {row} Error processing URL: {err}")

    finally:
        driver.quit()
        wb.save(main_excel_file)
        print("\n" + "=" * 60)
        print("All prices updated and issue reports successfully created!")


if __name__ == "__main__":
    main()
